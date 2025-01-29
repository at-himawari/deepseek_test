# app.py
from fastapi import FastAPI, HTTPException, File, UploadFile, Form
from pydantic import BaseModel
from typing import List, Optional
from transformers import AutoModelForCausalLM, AutoTokenizer, TextStreamer
from accelerate import disk_offload
import torch
from PIL import Image
import io
from docx import Document
import openpyxl

app = FastAPI(title="DeepSeek Multimodal API")

# データモデルの定義
class Message(BaseModel):
    role: str
    content: str

class RequestPayload(BaseModel):
    messages: List[Message] = []
    max_new_tokens: int = 4096
    temperature: float = 0.7

class ResponsePayload(BaseModel):
    generated_text: str

# モデルとトークナイザーのロード（アプリ起動時に一度だけ実行）
@app.on_event("startup")
def load_model():
    global model, tokenizer, streamer
    try:
        model = AutoModelForCausalLM.from_pretrained(
            "cyberagent/DeepSeek-R1-Distill-Qwen-32B-Japanese",
            torch_dtype="auto"
        )
        disk_offload(model=model, offload_dir="./.deepseek/offload")
        tokenizer = AutoTokenizer.from_pretrained("cyberagent/DeepSeek-R1-Distill-Qwen-32B-Japanese")
        streamer = TextStreamer(tokenizer, skip_prompt=True, skip_special_tokens=True)
    except Exception as e:
        print(f"Error loading model: {e}")

# ユーティリティ関数: 画像キャプション生成（仮の処理）
def generate_image_caption(pil_image: Image.Image) -> str:
    # ここに実際の画像キャプション生成モデルを統合してください。
    # 現在はサンプルのキャプションを返します。
    return "これはサンプルのキャプションです。"

# ユーティリティ関数: Officeファイルからテキストを抽出
def extract_text_from_office(file: UploadFile) -> str:
    try:
        if file.content_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' or file.filename.endswith('.docx'):
            document = Document(io.BytesIO(file.file.read()))
            text = "\n".join([para.text for para in document.paragraphs])
            return text
        elif file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or file.filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file.file.read()), read_only=True, data_only=True)
            text = ""
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
            return text
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file.content_type}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file {file.filename}: {str(e)}")

# マルチモーダル生成エンドポイント
@app.post("/multimodal_generate", response_model=ResponsePayload)
async def multimodal_generate(
    messages: Optional[str] = Form(None),
    max_new_tokens: Optional[int] = Form(4096),
    temperature: Optional[float] = Form(0.7),
    images: Optional[List[UploadFile]] = File(None),
    office_files: Optional[List[UploadFile]] = File(None)
):
    try:
        combined_text = ""

        # テキストメッセージの処理
        if messages:
            try:
                messages_json = Message.parse_raw(messages) if isinstance(messages, str) else messages
                # Pydanticモデルに基づいてパース
                # messagesはJSON文字列として送信される必要があります
                import json
                messages_list = json.loads(messages)
                input_messages = [Message(**msg) for msg in messages_list]
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Invalid messages format: {str(e)}")

            if input_messages:
                input_ids = tokenizer.apply_chat_template(
                    input_messages,
                    add_generation_prompt=True,
                    return_tensors="pt"
                ).to(model.device)

                output_ids = model.generate(
                    input_ids,
                    max_new_tokens=max_new_tokens,
                    temperature=temperature,
                    streamer=streamer
                )

                generated_text = tokenizer.decode(output_ids[0], skip_special_tokens=True)
                combined_text += generated_text + "\n"

        # 画像ファイルの処理
        if images:
            for image in images:
                if image.content_type.startswith('image/'):
                    image_data = await image.read()
                    pil_image = Image.open(io.BytesIO(image_data)).convert("RGB")
                    caption = generate_image_caption(pil_image)
                    combined_text += f"画像キャプション ({image.filename}): {caption}\n"
                else:
                    raise HTTPException(status_code=400, detail=f"Unsupported image type: {image.content_type}")

        # Officeファイルの処理
        if office_files:
            for office_file in office_files:
                text = extract_text_from_office(office_file)
                combined_text += f"ファイル内容 ({office_file.filename}):\n{text}\n"

        if not combined_text.strip():
            raise HTTPException(status_code=400, detail="No input provided.")

        return ResponsePayload(generated_text=combined_text.strip())

    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
